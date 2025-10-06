from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from app.core.supabase import get_supabase_admin_client
from app.core.config import settings
from typing import List, Dict, Any
from io import BytesIO
import uuid


class OutputFormatter:
    def __init__(self):
        self.supabase = get_supabase_admin_client()

    def generate_excel(
        self,
        data: List[Dict[str, Any]],
        columns: List[Dict[str, str]],
        job_id: str
    ) -> str:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        if columns:
            column_names = [col.get("name", f"Column{i}") for i, col in enumerate(columns)]
        elif data and len(data) > 0:
            column_names = list(data[0].keys())
        else:
            column_names = ["Data"]

        for col_idx, col_name in enumerate(column_names, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        if isinstance(data, list):
            for row_idx, row_data in enumerate(data, start=2):
                if isinstance(row_data, dict):
                    for col_idx, col_name in enumerate(column_names, start=1):
                        value = row_data.get(col_name, "")
                        sheet.cell(row=row_idx, column=col_idx, value=str(value))
                else:
                    sheet.cell(row=row_idx, column=1, value=str(row_data))

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        file_id = str(uuid.uuid4())
        storage_path = f"results/{job_id}/{file_id}.xlsx"

        self.supabase.storage.from_(settings.STORAGE_BUCKET_NAME).upload(
            storage_path,
            excel_buffer.getvalue(),
            file_options={"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        )

        file_url = self.supabase.storage.from_(settings.STORAGE_BUCKET_NAME).get_public_url(storage_path)

        return file_url

    def format_text_output(
        self,
        data: Any,
        template: str = None
    ) -> str:
        if template:
            try:
                return template.format(**data) if isinstance(data, dict) else str(data)
            except:
                pass

        if isinstance(data, dict):
            lines = []
            for key, value in data.items():
                formatted_key = key.replace("_", " ").title()
                lines.append(f"{formatted_key}: {value}")
            return "\n".join(lines)
        elif isinstance(data, list):
            return "\n".join([str(item) for item in data])
        else:
            return str(data)

    def validate_json_schema(
        self,
        data: Any,
        schema: Dict[str, Any]
    ) -> tuple[bool, List[str]]:
        errors = []

        if not schema or not schema.get("fields"):
            return True, []

        if not isinstance(data, dict):
            errors.append("Output must be a JSON object")
            return False, errors

        required_fields = [
            field["name"] for field in schema.get("fields", [])
            if field.get("required", False)
        ]

        for field in required_fields:
            if field not in data or data[field] is None:
                errors.append(f"Required field '{field}' is missing")

        for field_def in schema.get("fields", []):
            field_name = field_def.get("name")
            field_type = field_def.get("type", "string")

            if field_name in data:
                value = data[field_name]
                if not self._validate_field_type(value, field_type):
                    errors.append(f"Field '{field_name}' has invalid type. Expected {field_type}")

        return len(errors) == 0, errors

    def _validate_field_type(self, value: Any, expected_type: str) -> bool:
        if value is None:
            return True

        type_validators = {
            "string": lambda v: isinstance(v, str),
            "number": lambda v: isinstance(v, (int, float)),
            "integer": lambda v: isinstance(v, int),
            "boolean": lambda v: isinstance(v, bool),
            "array": lambda v: isinstance(v, list),
            "object": lambda v: isinstance(v, dict),
        }

        validator = type_validators.get(expected_type)
        return validator(value) if validator else True


output_formatter = OutputFormatter()
