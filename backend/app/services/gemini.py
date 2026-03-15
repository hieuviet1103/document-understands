"""
Gemini Document Understanding Service
======================================
Uses the new ``google-genai`` SDK (v1.x) with:

- **gemini-2.5-flash** (default; override via GEMINI_MODEL)
- **Files API** – documents > 20 MB are uploaded first then referenced
- **Native document understanding** – PDF / image bytes sent directly; no
  pre-processing needed
- **Structured output** – JSON mode via ``response_schema`` for guaranteed
  parseable responses
- **Synchronous API** – compatible with Celery workers (no event-loop needed)

Supported MIME types (native Gemini Document Understanding)
-----------------------------------------------------------
PDF, PNG, JPEG, WEBP, GIF, HEIC/HEIF, plain-text, HTML, CSS, JS, Python,
Markdown, CSV, RTF, DOCX (converted internally by Gemini)

For .xlsx/.xls we convert the spreadsheet to text (openpyxl) and send as text/plain.
"""

from __future__ import annotations

import copy
import io
import json
import logging
import time
from typing import Any, Dict, List, Optional

from google import genai
from google.genai import types
from openpyxl import load_workbook

from app.core.config import settings

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# MIME sets
# ─────────────────────────────────────────────────────────────────────────────

# MIME types that Gemini natively understands as *documents* (best quality)
_NATIVE_DOCUMENT_MIMES: frozenset[str] = frozenset({
    "application/pdf",
    # Images
    "image/png", "image/jpeg", "image/jpg",
    "image/webp", "image/gif", "image/heic", "image/heif",
    # Plain text variants
    "text/plain", "text/html", "text/css",
    "text/javascript", "application/javascript",
    "text/x-python", "text/x-script.python",
    "text/markdown", "text/csv",
    "application/rtf", "text/rtf",
    # Office (Gemini converts internally)
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",  # .docx
})

# Spreadsheet MIMEs not supported by Gemini; we convert to text before sending
_SPREADSHEET_MIMES: frozenset[str] = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # .xlsx
    "application/vnd.ms-excel",  # .xls
})


def _spreadsheet_to_text(file_content: bytes, mime_type: str) -> str:
    """Convert Excel (.xlsx) to plain text for Gemini. .xls is not supported."""
    if "spreadsheetml.sheet" not in mime_type:
        return "(Legacy .xls format is not supported. Please save as .xlsx.)"
    try:
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        parts: List[str] = []
        for sheet in wb.worksheets:
            parts.append(f"=== Sheet: {sheet.title} ===")
            for row in sheet.iter_rows(values_only=True):
                line = "\t".join(
                    str(cell) if cell is not None else ""
                    for cell in row
                )
                if line.strip():
                    parts.append(line)
            parts.append("")
        wb.close()
        return "\n".join(parts).strip() or "(Empty spreadsheet)"
    except Exception as e:
        logger.warning("Spreadsheet conversion failed: %s", e)
        return f"(Could not read spreadsheet: {e})"


class GeminiService:
    """High-level wrapper around the Gemini Document Understanding API."""

    def __init__(self) -> None:
        self.client = genai.Client(api_key=settings.GEMINI_API_KEY)
        self.model = settings.GEMINI_MODEL
        self._files_threshold = settings.GEMINI_FILES_THRESHOLD
        logger.info("GeminiService initialised — model=%s", self.model)

    # ──────────────────────────────────────────────────────────────────────────
    # Public entry point
    # ──────────────────────────────────────────────────────────────────────────

    def process_document(
        self,
        file_content: bytes,
        mime_type: str,
        output_format: str,
        schema: Optional[Dict[str, Any]] = None,
        custom_instructions: str = "",
        file_name: str = "document",
    ) -> Dict[str, Any]:
        """
        Analyse *file_content* with Gemini and return extracted data.

        Returns
        -------
        {
            "success": bool,
            "output":  str | dict | list,   # parsed result
            "tokens_used": int,
            "model": str,
            "error": str  # only when success=False
        }
        """
        try:
            doc_part = self._build_document_part(file_content, mime_type, file_name)
            prompt_part, gen_config = self._build_prompt_and_config(
                output_format, schema, custom_instructions
            )

            response = self.client.models.generate_content(
                model=self.model,
                contents=[doc_part, prompt_part],
                config=gen_config,
            )

            tokens = 0
            if response.usage_metadata:
                tokens = response.usage_metadata.total_token_count or 0

            output = self._parse_response(response.text, output_format)

            return {
                "success": True,
                "output": output,
                "tokens_used": tokens,
                "model": self.model,
            }

        except Exception as exc:
            logger.exception("Gemini processing error: %s", exc)
            return {
                "success": False,
                "error": str(exc),
                "tokens_used": 0,
                "model": self.model,
            }

    # ──────────────────────────────────────────────────────────────────────────
    # Document part construction
    # ──────────────────────────────────────────────────────────────────────────

    def _build_document_part(
        self, file_content: bytes, mime_type: str, file_name: str
    ) -> types.Part:
        """Return a ``types.Part`` for the document — inline or via Files API."""

        # Normalise mime type
        mime_type = mime_type.split(";")[0].strip().lower()

        # Gemini does not accept .xlsx/.xls; convert spreadsheet to text
        if mime_type in _SPREADSHEET_MIMES:
            text = _spreadsheet_to_text(file_content, mime_type)
            return types.Part.from_text(
                text="Document content (converted from Excel spreadsheet):\n\n" + text
            )

        if len(file_content) > self._files_threshold:
            # Large file → upload via Files API (survives up to 48 h)
            return self._upload_to_files_api(file_content, mime_type, file_name)

        # Small file → inline bytes (fastest, no quota impact)
        return types.Part.from_bytes(data=file_content, mime_type=mime_type)

    def _upload_to_files_api(
        self, file_content: bytes, mime_type: str, file_name: str
    ) -> types.Part:
        """Upload document to Files API and return a reference Part."""
        logger.info(
            "Uploading %d bytes to Files API (mime=%s)", len(file_content), mime_type
        )
        file_obj = self.client.files.upload(
            file=io.BytesIO(file_content),
            config=types.UploadFileConfig(
                name=file_name,
                mime_type=mime_type,
                display_name=file_name,
            ),
        )

        # Wait until file is ACTIVE (usually instant, but poll for safety)
        max_wait, waited = 30, 0
        while file_obj.state.name == "PROCESSING" and waited < max_wait:
            time.sleep(2)
            waited += 2
            file_obj = self.client.files.get(name=file_obj.name)

        if file_obj.state.name != "ACTIVE":
            raise RuntimeError(
                f"Files API upload failed — state={file_obj.state.name}"
            )

        logger.info("Files API upload complete: %s", file_obj.uri)
        return types.Part.from_uri(file_uri=file_obj.uri, mime_type=mime_type)

    # ──────────────────────────────────────────────────────────────────────────
    # Prompt + generation config builders
    # ──────────────────────────────────────────────────────────────────────────

    def _build_prompt_and_config(
        self,
        output_format: str,
        schema: Optional[Dict[str, Any]],
        custom_instructions: str,
    ) -> tuple[types.Part, types.GenerateContentConfig]:
        """Return (prompt_part, generation_config)."""

        base = (
            "You are an expert document analyst with deep knowledge of document "
            "understanding. Carefully read the entire document — including all "
            "pages, tables, figures, and footnotes — then extract the requested "
            "information with high accuracy.\n\n"
        )

        if custom_instructions:
            base += f"Additional Instructions:\n{custom_instructions}\n\n"

        if output_format == "text":
            prompt_text, gen_config = self._text_prompt_and_config(base, schema)
        elif output_format == "json":
            prompt_text, gen_config = self._json_prompt_and_config(base, schema)
        elif output_format == "excel":
            prompt_text, gen_config = self._excel_prompt_and_config(base, schema)
        else:
            prompt_text = base + "Extract all relevant information from the document."
            gen_config = types.GenerateContentConfig(
                temperature=0.1,
                max_output_tokens=8192,
            )

        return types.Part.from_text(text=prompt_text), gen_config

    # ── Text format ───────────────────────────────────────────────────────────

    def _text_prompt_and_config(
        self, base: str, schema: Optional[Dict[str, Any]]
    ) -> tuple[str, types.GenerateContentConfig]:
        if schema and schema.get("template"):
            template = schema["template"]
            prompt = (
                base
                + "Fill in the following template using information extracted from the document.\n"
                + "Replace every placeholder with the actual value found in the document.\n"
                + "Write 'N/A' if a value cannot be found.\n\n"
                + "TEMPLATE:\n"
                + template
            )
        else:
            prompt = (
                base
                + "Produce a clear, well-structured summary of the document.\n"
                + "Include: key facts, dates, names, amounts, decisions, and any action items."
            )

        config = types.GenerateContentConfig(
            temperature=0.2,
            max_output_tokens=8192,
        )
        return prompt, config

    # ── JSON format ───────────────────────────────────────────────────────────

    def _json_prompt_and_config(
        self, base: str, schema: Optional[Dict[str, Any]]
    ) -> tuple[str, types.GenerateContentConfig]:
        if schema and schema.get("fields"):
            fields: List[Dict] = schema["fields"]
            field_lines = self._describe_fields(fields)
            keys_hint = self._exact_keys_hint(fields)
            prompt = (
                base
                + "Extract the following fields from the document and return them as "
                + "a single valid JSON object. Support nested objects and arrays of objects.\n\n"
                + "FIELDS (nested structure; objects can have child fields, arrays contain items of the described shape):\n"
                + field_lines
                + "\n\nRules:\n"
                + "- Return ONLY the JSON object (no markdown fences, no commentary)\n"
                + "- The response must be a single JSON object with the top-level keys as listed above (e.g. \"users\"), not a bare array\n"
            )
            if keys_hint:
                prompt += f"- Use exactly these property names in the output: {keys_hint}\n"
            prompt += (
                "- Use null for fields that cannot be found\n"
                + "- Preserve nesting: objects with child fields, arrays of objects with their fields\n"
                + "- Ensure correct data types (string / number / integer / boolean / array / object)\n"
            )
            # Use response_json_schema (JSON Schema dict) so Gemini validates structure
            # without Pydantic Schema model rejecting our keys (e.g. no "name"/"children").
            # Deep copy + strip forbidden keys so the SDK never sees template-only fields.
            built = copy.deepcopy(self._build_response_schema(fields))
            response_json_schema = GeminiService._strip_forbidden_keys(built)
            config = types.GenerateContentConfig(
                temperature=0.0,
                max_output_tokens=8192,
                response_mime_type="application/json",
                response_json_schema=response_json_schema,
            )
        else:
            prompt = (
                base
                + "Extract all meaningful information from the document and return it "
                + "as a well-structured JSON object.\n"
                + "Return ONLY the JSON object — no markdown, no extra text.\n"
                + "Preserve nested structures (tables → arrays, sections → objects)."
            )
            config = types.GenerateContentConfig(
                temperature=0.0,
                max_output_tokens=8192,
                response_mime_type="application/json",
            )

        return prompt, config

    # ── Excel / tabular format ────────────────────────────────────────────────

    def _excel_prompt_and_config(
        self, base: str, schema: Optional[Dict[str, Any]]
    ) -> tuple[str, types.GenerateContentConfig]:
        if schema and schema.get("columns"):
            columns: List[Dict] = schema["columns"]
            col_names = [c.get("name", "") for c in columns]
            col_desc = "\n".join(
                f"  - {c.get('name')}: {c.get('description', '')}" for c in columns
            )
            example_obj = "{" + ", ".join(f'"{n}": "value"' for n in col_names) + "}"
            prompt = (
                base
                + "Extract tabular data from the document and return a JSON array "
                + "of row objects.\n\n"
                + "COLUMNS:\n"
                + col_desc
                + "\n\nRules:\n"
                + f"- Use EXACTLY these field names: {col_names}\n"
                + "- Each element of the array represents one row\n"
                + "- Use null for empty cells\n"
                + "- Return ONLY the JSON array (no markdown, no commentary)\n"
                + f"\nExample format:\n[{example_obj}, ...]"
            )
            response_schema = {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        c.get("name"): {"type": "string", "nullable": True}
                        for c in columns
                    },
                },
            }
            config = types.GenerateContentConfig(
                temperature=0.0,
                max_output_tokens=8192,
                response_mime_type="application/json",
                response_schema=response_schema,
            )
        else:
            prompt = (
                base
                + "Extract all tabular/row data from the document.\n"
                + "Return a JSON array where each element is a row object with "
                + "consistent key names.\n"
                + "Return ONLY the JSON array — no markdown, no extra text."
            )
            config = types.GenerateContentConfig(
                temperature=0.0,
                max_output_tokens=8192,
                response_mime_type="application/json",
            )

        return prompt, config

    # ──────────────────────────────────────────────────────────────────────────
    # Helpers
    # ──────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _exact_keys_hint(fields: List[Dict]) -> str:
        """Build a one-line hint of exact key names for the prompt."""
        root_keys = [f.get("name", "") for f in fields if f.get("name")]
        if not root_keys:
            return ""
        parts = [f"Root object must have key(s): {', '.join(repr(k) for k in root_keys)}"]
        for f in fields:
            if f.get("type") == "array" and isinstance(f.get("items"), dict):
                items = f.get("items", {})
                children = items.get("children") or items.get("properties")
                if isinstance(children, list):
                    child_names = [c.get("name", "") for c in children if c.get("name")]
                    if child_names:
                        parts.append(
                            f"each element of {f.get('name')!r} must have key(s): "
                            + ", ".join(repr(n) for n in child_names)
                        )
        return " ".join(parts) + "."

    @staticmethod
    def _describe_fields(fields: List[Dict], indent: str = "  ") -> str:
        """Describe fields for the prompt; supports nested object/array (children, items)."""
        lines: List[str] = []
        for f in fields:
            name = f.get("name", "")
            ftype = f.get("type", "string")
            desc = f.get("description", "")
            line = f"{indent}- {name} ({ftype})"
            if desc:
                line += f": {desc}"
            lines.append(line)
            if ftype == "object":
                children = f.get("children") or f.get("properties")
                if isinstance(children, list) and children:
                    lines.append(f"{indent}  (object with fields:)")
                    lines.append(GeminiService._describe_fields(children, indent + "    "))
                elif isinstance(children, dict):
                    lines.append(f"{indent}  properties: {json.dumps(children)}")
            elif ftype == "array":
                items = f.get("items")
                if isinstance(items, dict) and items.get("name") is not None:
                    lines.append(f"{indent}  (array of:)")
                    lines.append(GeminiService._describe_fields([items], indent + "    "))
                elif isinstance(items, dict):
                    lines.append(f"{indent}  items: {json.dumps(items)}")
        return "\n".join(lines)

    # Keys allowed in response_schema / response_json_schema (SDK may validate; no extra keys)
    _SCHEMA_ALLOWED_KEYS = frozenset({"type", "nullable", "properties", "items", "description"})
    _SCHEMA_FORBIDDEN_KEYS = frozenset({"name", "children"})  # template-only; must never reach SDK

    @staticmethod
    def _sanitize_json_schema(schema: Dict[str, Any]) -> Dict[str, Any]:
        """Return a copy with only JSON Schema keys; strip name/children so SDK never sees them."""
        if not isinstance(schema, dict):
            return schema
        out: Dict[str, Any] = {}
        for k, v in schema.items():
            if k in GeminiService._SCHEMA_FORBIDDEN_KEYS or k not in GeminiService._SCHEMA_ALLOWED_KEYS:
                continue
            if k == "properties" and isinstance(v, dict):
                out[k] = {pk: GeminiService._sanitize_json_schema(pv) for pk, pv in v.items()}
            elif k == "items" and isinstance(v, dict):
                out[k] = GeminiService._sanitize_json_schema(v)
            else:
                out[k] = v
        return out

    @staticmethod
    def _strip_forbidden_keys(obj: Any) -> Any:
        """Recursively remove 'name' and 'children' from dicts (defense in depth)."""
        if not isinstance(obj, dict):
            return obj
        out: Dict[str, Any] = {}
        for k, v in obj.items():
            if k in GeminiService._SCHEMA_FORBIDDEN_KEYS:
                continue
            if isinstance(v, dict):
                out[k] = GeminiService._strip_forbidden_keys(v)
            elif isinstance(v, list):
                out[k] = [GeminiService._strip_forbidden_keys(i) for i in v]
            else:
                out[k] = v
        return out

    @staticmethod
    def _field_to_json_schema(field: Dict[str, Any]) -> Dict[str, Any]:
        """Turn a single field definition into a JSON Schema fragment (for Gemini).
        Returns only allowed keys (type, nullable, properties, items, description).
        """
        ftype = field.get("type", "string")
        if ftype not in ("string", "number", "integer", "boolean", "array", "object"):
            ftype = "string"
        prop: Dict[str, Any] = {"type": ftype, "nullable": True}
        if ftype == "object":
            children = field.get("children") or field.get("properties")
            if isinstance(children, list):
                prop["properties"] = GeminiService._build_response_schema(children).get("properties", {})
            elif isinstance(children, dict):
                prop["properties"] = {k: GeminiService._sanitize_json_schema(v) for k, v in children.items()}
            else:
                prop["properties"] = {}
        elif ftype == "array":
            items = field.get("items")
            if isinstance(items, dict):
                # Build item schema from template; sanitize so SDK never sees name/children
                prop["items"] = GeminiService._sanitize_json_schema(
                    GeminiService._field_to_json_schema(items)
                )
            else:
                prop["items"] = {"type": "string", "nullable": True}
        return GeminiService._sanitize_json_schema(prop)

    @staticmethod
    def _build_response_schema(fields: List[Dict]) -> Dict[str, Any]:
        """Convert field definitions (with optional children/items) to JSON Schema for Gemini."""
        properties: Dict[str, Any] = {}
        for f in fields:
            name = f.get("name", "")
            if not name:
                continue
            properties[name] = GeminiService._field_to_json_schema(f)
        root = {"type": "object", "properties": properties}
        return GeminiService._sanitize_json_schema(root)

    @staticmethod
    def _parse_response(response_text: str, output_format: str) -> Any:
        """Strip markdown fences and parse JSON when required."""
        text = response_text.strip()

        # Remove possible markdown code fences
        if text.startswith("```json"):
            text = text[7:]
        elif text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()

        if output_format == "text":
            return text

        # JSON / Excel → parse
        try:
            return json.loads(text)
        except json.JSONDecodeError as exc:
            logger.warning("JSON parse error: %s — returning raw text", exc)
            return {"raw_output": text, "parse_error": str(exc)}


# ── Module-level singleton ────────────────────────────────────────────────────
gemini_service = GeminiService()
